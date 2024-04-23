# user guide: put the file "testResult.txt" in the same folder with the program, then run the program, 
# then you will get "analysisResult.txt" in the same folder
# dataList = [-1000,1,2,13,14,100,101,1000]
import openpyxl
import os
maxR = 6# the maximum of R that we can tolerate
maxTurn = 1000
import matplotlib.pyplot as plt




r1 = 0

def takeData(stri):
    dataList = []
    f=open('testResult.txt', encoding='gbk')
    txt=[]
    for line in f:
        txt.append(line.strip())
#print(txt)
#print(len(txt))
    for i in range(0,len(txt)):
        index = txt[i].find(stri)
        if(index != -1):
        # print(index)
            end = 6
            for j in range(0,7):
            # print(txt[i][index + len(stri) + 2 + j])
            # print(txt[i][index + len(stri) + 2 + j] == ",")
                if(txt[i][index + len(stri) + 2 + j] == ","):
                    end = j-1
                    break
            s = txt[i][index +len(stri) + 2:index + len(stri) + 2 + end]
        # print(s)
            dataList.append(eval(s))
    dataList.sort()
    #print(dataList)
    n = len(dataList)
    return dataList, n


# these paras are only can be changed by changing paras in function
def UniformNum(r):
    # the num compared with results in different r 
    return 1/12/r**3

def findIndex(dataList,low , high ,num ):
    # find the index such dataList[index] <= num < dataList[index +1] 
    #-1 when num < dataList[0]
    # print(low,high,num)
    index = int((low + high)/2)
    #print(type(index))
    #print(index,low, high, num)
    if(num >= dataList[index]):
        if(index == high):
            return high
        if(num < dataList[index + 1]):
            return index
        else:
            return findIndex(dataList=dataList,low = index + 1, high= high, num = num)
    else:
        if(index == 0):
            return -1
        else:
            return findIndex( dataList=dataList,low = low, high = index - 1 ,num = num)
    
# s = findIndex(dataList = dataList, num = 6.4)
# print(s)
def FindCentralList(dataList ,r ,step ,maxTurn = 1000):
    # find the representatives of dataLiST with length r;
    turncout = 0
    centralList = initCentralList(dataList,r)
    # print(centralList)
    
    while(CheckEnd(dataList,centralList, r) != True and turncout < maxTurn):
        centralList = adjust(dataList,centralList,r,step)
        turncout +=1
        # print(centralList)
        
    return centralList
def initCentralList(dataList,r):
    n = len(dataList)
    #initial a list with [-10000,c1,c2,''',cr ,10000] ci = l[i]
    l = []
    n = len(dataList)
    l.append(-10000)
    for i in range(1,r+1):
        l.append((dataList[n-1] - dataList[0])/(r+1)*i + dataList[0])
    l.append(10000)
    return l
def findSi(dataList ,lowNum , highNum ):
    n = len(dataList)
    # find the index of dataList whose values between low and high
    sm = findIndex(dataList,low = 0, high = n - 1,num = lowNum) + 1
    sn = findIndex(dataList,low = 0, high = n - 1,num = highNum) 
    return sm , sn
def adjust(dataList,centralList, r,step):
    n = len(dataList)
    for i in range(1,r+1):
        fi = 0
        scount = 0
        Sm, Sn = findSi( dataList,lowNum = (centralList[i] + centralList[i-1])/2, 
        highNum = (centralList[i] + centralList[i+1])/2)
        #print(Sm, Sn)
        for j in range(Sm,Sn + 1):
            fi += (dataList[j] - centralList[i])/n
            scount += 1/n
        # print("fi")
        # print(fi)
        
        if(i != 1 and Sm != 0 and Sm <= Sn):
            Pdown =(centralList[i] - centralList[i-1])/4*1/n/(dataList[Sm] - dataList[Sm -1])
        else:
            Pdown = 0
        if(i != r and Sn != n-1 and Sm <= Sn):
            # print(i,r)
            Pup = (centralList[i+1] - centralList[i])/4*1/n/(dataList[Sn + 1] - dataList[Sn])
        else:
            Pup = 0
        dfi = Pup + Pdown - scount
        if(dfi == 0): dfi = -1
        # print("dfi")
        # print(dfi)
        if(fi != 0):
            centralList[i] = max( centralList[i] - fi/dfi/abs(fi/dfi)*step, centralList[i-1]+0.01)
        else:
            centralList[i] = centralList[i]
        centralList[i] = max(centralList[i],dataList[0] + step/1000)
        centralList[i] = min(centralList[i],dataList[n-1] - step/1000)
    return centralList

def CheckEnd(dataList,centralList, r):
    allrightposition = True
    n = len(dataList)
    for i in range(1,r+1):
            xi = 0
            sum = 0
            count = 0
            Sm, Sn = findSi(dataList, lowNum = (centralList[i] + centralList[i-1])/2, 
            highNum = (centralList[i] + centralList[i+1])/2)
            for j in range(Sm,Sn + 1):
                sum += dataList[j]
                count += 1
            if(count != 0):
                xi =  sum/ count
                indexx = findIndex(dataList,low = 0,high = n - 1,num = xi)
                indexi = findIndex(dataList,low = 0,high = n - 1,num = centralList[i])
                if(indexi != indexx):
                    allrightposition = False
            else:
                allrightposition = False
    return allrightposition

def FindAndFinalAdujust(dataList,r,step, maxTurn):
    n = len(dataList)
    centralList = FindCentralList(dataList,r = r,step = step,maxTurn= maxTurn)
    var = 0
    for i in range(1,r+1):
        sum = 0
        scount = 0
        Sm, Sn = findSi( dataList,lowNum = (centralList[i] + centralList[i-1])/2, 
        highNum = (centralList[i] + centralList[i+1])/2)
        # print(Sm, Sn)
        for j in range(Sm,Sn + 1):
            sum += dataList[j]
            scount += 1
        if(scount == 0): 
            continue
        centralList[i] = sum/ scount
        for j in range(Sm,Sn + 1):
           var += (dataList[j] - centralList[i])**2/n
    return centralList, var

def DecideR(dataList,step ,maxR = maxR,maxTurn = 1000):
    minRevisedVar = 0
    relatedR  = 1
    relatedCList = []
    revisedVarList = []
    allClist = []
    for R in range(1,maxR+1):
        cList, var = FindAndFinalAdujust(dataList,r = R,step = step,maxTurn= maxTurn)
        for counting in range(len(cList)):
            cList[counting] = rounds(cList[counting])
        revisedVar = rounds(var/UniformNum(R))
        revisedVarList.append(revisedVar)
        allClist.append(cList[1:len(cList)-1])
        if(R == 1):
            relatedR = R
            minRevisedVar =rounds( revisedVar)
            relatedCList = cList
        if(revisedVar < minRevisedVar):
            relatedR = R
            minRevisedVar = revisedVar
            relatedCList = cList
    return relatedCList, relatedR, minRevisedVar ,revisedVarList, allClist
def rounds(n):
    return int( 10000*n)/10000


stris = ["blinkTime","errorRadius","judgeTime","judgeTimeOther"]
result =""
data = openpyxl.Workbook() # 新建工作簿
for i in range(len(stris)):
    dataList, n = takeData(stris[i])
    #print(dataList,n)
    stepRevised = 0.01
    if(stris[i] == "blinkTime"):
        stepRevised = 0.001
    if(stris[i] == "errorRadius"):
        stepRevised = 1
    if(stris[i] == "judgeTime" or stris[i] == "judgeTimeOther"):
        stepRevised = 0.01
    li,r,var,varlist,allClist =DecideR(dataList,step = stepRevised, maxR= maxR,maxTurn = maxTurn)
    #print(DecideR(dataList,step = stepRevised, maxR= maxR,maxTurn = maxTurn))
    dic = {"name of parameter": stris[i],
    'data List':dataList,"related Representatives":li[1:len(li)-1],"best Number Of Representatives":r,
    "the Number of Representatives have been tried":maxR,
    "the max Number of Turns(to avoid endless work)":maxTurn,
    "revised Variance":var,
    "Variance List With Different Choices Of NumOfRepresentatives":varlist,"all centralList":allClist}
    result += str(dic) +"\n"

    
    data.create_sheet(stris[i]) # 添加页
    #table = data.get_sheet_by_name('Sheet1') # 获得指定名称页
    table =data[stris[i]] # 获得当前活跃的工作页，默认为第一个工作页
    print(table)
    table.cell(1,1,'name of parameter') # 行，列，值 这里是从1开始计数的
    table.cell(1,2,stris[i])
    table.cell(2,1,'best Number Of Representatives') 
    table.cell(2,2,r) 
    table.cell(3,1,"the Number of Representatives have been tried")
    table.cell(3,2,maxR)
    table.cell(4,1,"the max Number of Turns(to avoid endless work)")
    table.cell(4,2,maxTurn)
    table.cell(5,1,"min revised Variance")
    table.cell(5,2,var)
    table.cell(1,4,"data(from low to high)\|/")
    for cou in range(len(dataList)):
        table.cell(cou + 2,4, dataList[cou])
    table.cell(1,5,"the Number of Representatives-->")
    table.cell(10,5, "related revised varience")
    table.cell(3,5, "chosen num")
    for cou in range(len(allClist)):
        table.cell(1,cou + 6, cou + 1)
        table.cell(10, cou + 6, varlist[cou])
        for j in range(len(allClist[cou])):
            table.cell(j + 2, cou + 6, allClist[cou][j] )
    fig = plt.figure()
    x=[0]*len(dataList)
    y = dataList
    m = min(dataList)
    plt.scatter(x,  y, c= "red", label = "basic data",alpha = 1/6)
    for u in range(maxR):
        x = [u+1]*len(allClist[u])
        y = allClist[u]
        plt.scatter(x,y, c= "blue", label = "revised variance = "+ str(varlist[i]))
    
    plt.xlabel("number of representatives")
    plt.ylabel("value of parameters")
    plt.text(0,m - 1,"data")
    #绘制图片
    plt.savefig( stris[i]+'.png')#存储图片时给图片命名，注意要放在plt.show()前
    plt.show()
    #print(table.cell(1,2))
data.save('analysisResult.xlsx')

with open('analysisResult.txt', 'w') as f:
        f.write(result)




    



            

#print(FindCentralList(r = 2,step= 0.01 ))
#print(FindAndFinalAdujust(r=3, step = 0.01)) 
